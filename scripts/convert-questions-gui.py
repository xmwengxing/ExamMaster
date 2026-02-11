#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
题库格式转换工具 - GUI版本
提供图形界面的题库格式转换工具，支持Excel和Word文件转换为JSON格式

使用方法：
1. 安装依赖：pip install openpyxl python-docx pillow
2. 运行脚本：python scripts/convert-questions-gui.py

功能特性：
- 图形化界面，操作简单
- 支持Excel (.xlsx, .xls) 和 Word (.docx) 格式
- 实时进度显示
- 详细的错误提示
- 自动生成JSON格式输出
"""

import os
import re
import json
import sys
import threading
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# 检查依赖库
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from docx import Document
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False

try:
    from PIL import Image
    import io
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class QuestionBankConverterGUI:
    """题库转换工具GUI主类"""
    
    def __init__(self):
        """初始化GUI窗口"""
        self.window = tk.Tk()
        self.window.title("题库格式转换工具 v2.0")
        self.window.geometry("700x600")
        self.window.resizable(False, False)
        
        # 设置窗口图标（如果有的话）
        try:
            # self.window.iconbitmap('icon.ico')
            pass
        except:
            pass
        
        # 变量
        self.source_file_path = tk.StringVar()
        self.output_dir_path = tk.StringVar()
        self.file_format = tk.StringVar(value='excel')
        self.is_converting = False
        self.conversion_thread = None
        self.debug_mode = False  # 调试模式标志
        
        # 设置默认输出目录为当前目录
        self.output_dir_path.set(os.getcwd())
        
        # 创建UI组件
        self.setup_ui()
        
        # 检查依赖
        self.check_dependencies()
    
    def setup_ui(self):
        """设置UI布局"""
        # 主容器
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 标题
        title_label = ttk.Label(
            main_frame, 
            text="题库格式转换工具", 
            font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # 文件选择区域
        self.create_file_selection_section(main_frame)
        
        # 输出目录选择区域
        self.create_output_dir_section(main_frame)
        
        # 文件格式选择区域
        self.create_format_selection_section(main_frame)
        
        # 进度显示区域
        self.create_progress_section(main_frame)
        
        # 日志显示区域
        self.create_log_section(main_frame)
        
        # 按钮区域
        self.create_button_section(main_frame)
        
        # 状态栏
        self.create_status_bar()
    
    def create_file_selection_section(self, parent):
        """创建文件选择区域"""
        # 源文件标签
        file_label = ttk.Label(parent, text="源文件:")
        file_label.grid(row=1, column=0, sticky=tk.W, pady=5)
        
        # 源文件输入框
        self.file_entry = ttk.Entry(
            parent, 
            textvariable=self.source_file_path, 
            width=50
        )
        self.file_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        
        # 浏览按钮
        browse_button = ttk.Button(
            parent, 
            text="浏览...", 
            command=self.browse_file
        )
        browse_button.grid(row=1, column=2, pady=5)
    
    def create_output_dir_section(self, parent):
        """创建输出目录选择区域"""
        # 输出目录标签
        output_label = ttk.Label(parent, text="输出目录:")
        output_label.grid(row=2, column=0, sticky=tk.W, pady=5)
        
        # 输出目录输入框
        self.output_entry = ttk.Entry(
            parent, 
            textvariable=self.output_dir_path, 
            width=50
        )
        self.output_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        
        # 浏览按钮
        output_browse_button = ttk.Button(
            parent, 
            text="浏览...", 
            command=self.browse_output_dir
        )
        output_browse_button.grid(row=2, column=2, pady=5)
    
    def create_format_selection_section(self, parent):
        """创建文件格式选择区域"""
        # 格式标签
        format_label = ttk.Label(parent, text="源文件格式:")
        format_label.grid(row=3, column=0, sticky=tk.W, pady=5)
        
        # 格式选择框架
        format_frame = ttk.Frame(parent)
        format_frame.grid(row=3, column=1, sticky=tk.W, pady=5, padx=5)
        
        # Excel单选按钮
        excel_radio = ttk.Radiobutton(
            format_frame, 
            text="Excel (.xlsx, .xls)", 
            variable=self.file_format, 
            value='excel'
        )
        excel_radio.grid(row=0, column=0, padx=(0, 20))
        
        # Word单选按钮
        word_radio = ttk.Radiobutton(
            format_frame, 
            text="Word (.docx)", 
            variable=self.file_format, 
            value='word'
        )
        word_radio.grid(row=0, column=1)
    
    def create_progress_section(self, parent):
        """创建进度显示区域"""
        # 进度框架
        progress_frame = ttk.LabelFrame(parent, text="转换进度", padding="10")
        progress_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # 进度条
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            mode='determinate', 
            length=600
        )
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        # 进度标签
        self.progress_label = ttk.Label(progress_frame, text="就绪")
        self.progress_label.grid(row=1, column=0, sticky=tk.W)
    
    def create_log_section(self, parent):
        """创建日志显示区域"""
        # 日志框架
        log_frame = ttk.LabelFrame(parent, text="转换日志", padding="10")
        log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # 日志文本框（带滚动条）
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            width=70, 
            height=10, 
            wrap=tk.WORD,
            state='disabled'
        )
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    def create_button_section(self, parent):
        """创建按钮区域"""
        # 按钮框架
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=6, column=0, columnspan=3, pady=10)
        
        # 测试转换按钮
        self.test_button = ttk.Button(
            button_frame, 
            text="测试转换", 
            command=self.start_test_conversion,
            width=15
        )
        self.test_button.grid(row=0, column=0, padx=5)
        
        # 开始转换按钮
        self.convert_button = ttk.Button(
            button_frame, 
            text="开始转换", 
            command=self.start_conversion,
            width=15
        )
        self.convert_button.grid(row=0, column=1, padx=5)
        
        # 取消按钮
        self.cancel_button = ttk.Button(
            button_frame, 
            text="取消", 
            command=self.cancel_conversion,
            width=15,
            state='disabled'
        )
        self.cancel_button.grid(row=0, column=2, padx=5)
        
        # 打开输出目录按钮
        self.open_button = ttk.Button(
            button_frame, 
            text="打开输出目录", 
            command=self.open_output_directory,
            width=15,
            state='disabled'
        )
        self.open_button.grid(row=0, column=3, padx=5)
    
    def create_status_bar(self):
        """创建状态栏"""
        self.status_bar = ttk.Label(
            self.window, 
            text="就绪", 
            relief=tk.SUNKEN, 
            anchor=tk.W
        )
        self.status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
    
    def browse_file(self):
        """浏览并选择源文件"""
        filetypes = [
            ("所有支持的格式", "*.xlsx *.xls *.docx"),
            ("Excel文件", "*.xlsx *.xls"),
            ("Word文件", "*.docx"),
            ("所有文件", "*.*")
        ]
        
        filename = filedialog.askopenfilename(
            title="选择题库文件",
            filetypes=filetypes
        )
        
        if filename:
            self.source_file_path.set(filename)
            # 根据文件扩展名自动选择格式
            ext = Path(filename).suffix.lower()
            if ext in ['.xlsx', '.xls']:
                self.file_format.set('excel')
            elif ext == '.docx':
                self.file_format.set('word')
            
            self.log_message(f"已选择文件: {filename}")
    
    def browse_output_dir(self):
        """浏览并选择输出目录"""
        directory = filedialog.askdirectory(
            title="选择输出目录",
            initialdir=self.output_dir_path.get()
        )
        
        if directory:
            self.output_dir_path.set(directory)
            self.log_message(f"输出目录: {directory}")
    
    def check_dependencies(self):
        """检查依赖库"""
        missing_deps = []
        
        if not EXCEL_AVAILABLE:
            missing_deps.append("openpyxl (Excel支持)")
        
        if not WORD_AVAILABLE:
            missing_deps.append("python-docx (Word支持)")
        
        if not PIL_AVAILABLE:
            missing_deps.append("pillow (图片压缩支持)")
        
        if missing_deps:
            warning_msg = "警告: 缺少以下依赖库:\n\n"
            warning_msg += "\n".join(f"- {dep}" for dep in missing_deps)
            warning_msg += "\n\n安装命令:\n"
            warning_msg += "pip install openpyxl python-docx pillow"
            
            self.log_message(warning_msg)
            messagebox.showwarning("依赖检查", warning_msg)
    
    def log_message(self, message):
        """在日志区域显示消息"""
        def _log():
            self.log_text.config(state='normal')
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        
        # 确保在主线程中更新UI
        if threading.current_thread() == threading.main_thread():
            _log()
        else:
            self.window.after(0, _log)
    
    def update_progress(self, value, message=""):
        """更新进度条和进度标签"""
        def _update():
            self.progress_bar['value'] = value
            if message:
                self.progress_label.config(text=message)
        
        # 确保在主线程中更新UI
        if threading.current_thread() == threading.main_thread():
            _update()
        else:
            self.window.after(0, _update)
    
    def update_status(self, message):
        """更新状态栏"""
        def _update():
            self.status_bar.config(text=message)
        
        # 确保在主线程中更新UI
        if threading.current_thread() == threading.main_thread():
            _update()
        else:
            self.window.after(0, _update)
    
    def start_test_conversion(self):
        """开始测试转换（每种题型各转换5题）"""
        # 验证输入
        source_file = self.source_file_path.get()
        if not source_file:
            messagebox.showerror("错误", "请选择源文件")
            return
        
        if not os.path.exists(source_file):
            messagebox.showerror("错误", "源文件不存在")
            return
        
        # 检查文件格式和依赖
        file_format = self.file_format.get()
        if file_format == 'excel' and not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装openpyxl库，无法处理Excel文件\n\n安装命令: pip install openpyxl")
            return
        
        if file_format == 'word' and not WORD_AVAILABLE:
            messagebox.showerror("错误", "未安装python-docx库，无法处理Word文件\n\n安装命令: pip install python-docx")
            return
        
        # 禁用按钮
        self.test_button.config(state='disabled')
        self.convert_button.config(state='disabled')
        self.cancel_button.config(state='normal')
        self.is_converting = True
        
        # 清空日志
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        # 重置进度
        self.update_progress(0, "准备测试转换...")
        self.update_status("正在测试转换...")
        
        # 在新线程中执行测试转换
        self.conversion_thread = threading.Thread(
            target=self.test_convert_file,
            args=(source_file, file_format)
        )
        self.conversion_thread.start()
    
    def start_conversion(self):
        """开始转换"""
        # 验证输入
        source_file = self.source_file_path.get()
        if not source_file:
            messagebox.showerror("错误", "请选择源文件")
            return
        
        if not os.path.exists(source_file):
            messagebox.showerror("错误", "源文件不存在")
            return
        
        output_dir = self.output_dir_path.get()
        if not output_dir:
            messagebox.showerror("错误", "请选择输出目录")
            return
        
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                messagebox.showerror("错误", f"无法创建输出目录: {str(e)}")
                return
        
        # 检查文件格式和依赖
        file_format = self.file_format.get()
        if file_format == 'excel' and not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装openpyxl库，无法处理Excel文件\n\n安装命令: pip install openpyxl")
            return
        
        if file_format == 'word' and not WORD_AVAILABLE:
            messagebox.showerror("错误", "未安装python-docx库，无法处理Word文件\n\n安装命令: pip install python-docx")
            return
        
        # 禁用转换按钮，启用取消按钮
        self.test_button.config(state='disabled')
        self.convert_button.config(state='disabled')
        self.cancel_button.config(state='normal')
        self.open_button.config(state='disabled')
        self.is_converting = True
        
        # 清空日志
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        # 重置进度
        self.update_progress(0, "准备转换...")
        self.update_status("正在转换...")
        
        # 在新线程中执行转换
        self.conversion_thread = threading.Thread(
            target=self.convert_file,
            args=(source_file, output_dir, file_format)
        )
        self.conversion_thread.start()
    
    def cancel_conversion(self):
        """取消转换"""
        if messagebox.askyesno("确认", "确定要取消转换吗？"):
            self.is_converting = False
            self.log_message("用户取消了转换")
            self.update_status("已取消")
            self.reset_ui()
    
    def test_convert_file(self, source_file, file_format):
        """测试转换文件（每种题型各转换5题）"""
        try:
            # 启用调试模式
            self.debug_mode = True
            
            self.log_message(f"开始测试转换: {source_file}")
            self.log_message(f"文件格式: {file_format}")
            self.log_message("=" * 60)
            self.log_message("【调试模式已启用】将输出详细的题型识别过程")
            self.log_message("=" * 60)
            
            # 解析文件（不限制数量，后续按题型筛选）
            if file_format == 'excel':
                all_questions = self.parse_excel_file(source_file)
            elif file_format == 'word':
                all_questions = self.parse_word_file(source_file)
            else:
                raise ValueError(f"不支持的文件格式: {file_format}")
            
            # 关闭调试模式
            self.debug_mode = False
            
            self.log_message("=" * 60)
            self.log_message("【题型识别完成】")
            self.log_message("=" * 60)
            
            if not all_questions:
                # 转换失败，显示详细错误
                self.show_test_error("未能提取到题目", 
                    "可能的原因：\n" +
                    "1. 文件格式不正确\n" +
                    "2. 文件中没有有效的题目数据\n" +
                    "3. 题目格式不符合要求\n\n" +
                    "请检查文件内容是否包含题目、选项、答案等信息")
                return
            
            # 显示总题数
            self.log_message(f"文档解析完成，共 {len(all_questions)} 道题目")
            self.log_message("")
            
            # 按题型分类
            questions_by_type = {
                'single': [],
                'multiple': [],
                'judge': [],
                'fill': [],
                'essay': []
            }
            
            for q in all_questions:
                q_type = q.get('type', 'single')
                if q_type in questions_by_type:
                    questions_by_type[q_type].append(q)
                else:
                    # 未知题型，归类到single
                    self.log_message(f"警告: 发现未知题型 '{q_type}'，已归类为单选题")
                    questions_by_type['single'].append(q)
            
            # 显示题型统计（包括0题的题型）
            type_names = {
                'single': '单选题',
                'multiple': '多选题',
                'judge': '判断题',
                'fill': '填空题',
                'essay': '简答题'
            }
            
            self.log_message("题型统计：")
            for q_type in ['single', 'multiple', 'judge', 'fill', 'essay']:
                count = len(questions_by_type[q_type])
                self.log_message(f"  {type_names[q_type]}: {count} 题")
            self.log_message("")
            
            # 每种题型取5题用于预览
            test_questions = []
            self.log_message("预览题目（每种题型最多5题）：")
            
            for q_type in ['single', 'multiple', 'judge', 'fill', 'essay']:
                q_list = questions_by_type[q_type]
                count = min(5, len(q_list))
                if count > 0:
                    test_questions.extend(q_list[:count])
                    self.log_message(f"  {type_names[q_type]}: 预览 {count} 题")
            
            if not test_questions:
                self.show_test_error("未能提取到题目", 
                    "文件中没有找到有效的题目数据")
                return
            
            self.log_message(f"测试转换成功！共提取 {len(test_questions)} 道题目")
            self.update_progress(100, f"测试完成，成功转换 {len(test_questions)} 题")
            self.update_status("测试成功")
            
            # 显示预览窗口
            self.window.after(0, lambda: self.show_preview_window(test_questions))
            
        except FileNotFoundError as e:
            error_msg = "文件不存在"
            detail = f"无法找到文件: {source_file}\n\n请检查文件路径是否正确"
            self.show_test_error(error_msg, detail)
            
        except PermissionError as e:
            error_msg = "文件访问权限错误"
            detail = f"无法读取文件: {source_file}\n\n可能的原因：\n1. 文件正在被其他程序使用\n2. 没有读取权限\n\n请关闭文件后重试"
            self.show_test_error(error_msg, detail)
            
        except ValueError as e:
            error_msg = "数据格式错误"
            detail = f"{str(e)}\n\n请检查：\n1. 文件是否包含题目数据\n2. 题目格式是否正确\n3. 是否选择了正确的文件格式"
            self.show_test_error(error_msg, detail)
            
        except Exception as e:
            error_msg = "转换失败"
            detail = f"错误信息: {str(e)}\n\n请检查：\n1. 文件是否损坏\n2. 文件格式是否正确\n3. 是否安装了必要的依赖库"
            
            # 获取详细错误信息
            import traceback
            error_trace = traceback.format_exc()
            self.log_message(f"详细错误:\n{error_trace}")
            
            self.show_test_error(error_msg, detail)
        
        finally:
            self.window.after(0, self.reset_ui)
    
    def show_test_error(self, error_type, detail):
        """显示测试转换错误"""
        self.log_message(f"测试转换失败: {error_type}")
        self.log_message(detail)
        self.update_status("测试失败")
        
        def _show():
            messagebox.showerror(
                f"测试转换失败 - {error_type}",
                detail
            )
        
        self.window.after(0, _show)
    
    def clean_content_for_preview(self, content):
        """清理内容用于预览显示，将Base64图片替换为简单标识"""
        import re
        
        if not content:
            return content
        
        # 替换Base64图片为简单标识
        # 匹配 <img src='data:image/...;base64,...' />
        pattern = r"<img\s+src=['\"]data:image/[^;]+;base64,[^'\"]+['\"](?:\s+[^>]*)?\s*/?>"
        cleaned = re.sub(pattern, '[图片]', content)
        
        # 如果内容过长，截断显示
        if len(cleaned) > 500:
            cleaned = cleaned[:500] + '...'
        
        return cleaned
    
    def show_preview_window(self, questions):
        """显示预览窗口"""
        # 创建预览窗口
        preview_window = tk.Toplevel(self.window)
        preview_window.title(f"转换预览 - 共 {len(questions)} 题")
        preview_window.geometry("800x600")
        
        # 创建主框架
        main_frame = ttk.Frame(preview_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(
            main_frame,
            text=f"测试转换成功！共提取 {len(questions)} 道题目",
            font=("Arial", 12, "bold")
        )
        title_label.pack(pady=(0, 10))
        
        # 创建滚动文本框
        preview_text = scrolledtext.ScrolledText(
            main_frame,
            width=90,
            height=30,
            wrap=tk.WORD,
            font=("Consolas", 10)
        )
        preview_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 填充预览内容
        for idx, q in enumerate(questions, 1):
            preview_text.insert(tk.END, f"{'='*80}\n")
            preview_text.insert(tk.END, f"第 {idx} 题\n")
            preview_text.insert(tk.END, f"{'='*80}\n\n")
            
            preview_text.insert(tk.END, f"题型: {self.get_type_name(q['type'])}\n\n")
            
            # 清理题目内容中的Base64图片
            cleaned_content = self.clean_content_for_preview(q['content'])
            preview_text.insert(tk.END, f"题目: {cleaned_content}\n\n")
            
            if q.get('options'):
                preview_text.insert(tk.END, "选项:\n")
                for opt_idx, opt in enumerate(q['options']):
                    # 清理选项中的Base64图片
                    cleaned_opt = self.clean_content_for_preview(opt)
                    preview_text.insert(tk.END, f"  {chr(65 + opt_idx)}. {cleaned_opt}\n")
                preview_text.insert(tk.END, "\n")
            
            # 清理答案中的Base64图片
            cleaned_answer = self.clean_content_for_preview(str(q['answer']))
            preview_text.insert(tk.END, f"答案: {cleaned_answer}\n\n")
            
            if q.get('explanation'):
                # 清理解析中的Base64图片
                cleaned_explanation = self.clean_content_for_preview(q['explanation'])
                preview_text.insert(tk.END, f"解析: {cleaned_explanation}\n\n")
            
            preview_text.insert(tk.END, "\n")
        
        preview_text.config(state='disabled')
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=(10, 0))
        
        # 继续转换按钮
        continue_button = ttk.Button(
            button_frame,
            text="继续完整转换",
            command=lambda: [preview_window.destroy(), self.start_conversion()],
            width=20
        )
        continue_button.pack(side=tk.LEFT, padx=5)
        
        # 关闭按钮
        close_button = ttk.Button(
            button_frame,
            text="关闭",
            command=preview_window.destroy,
            width=20
        )
        close_button.pack(side=tk.LEFT, padx=5)
    
    def get_type_name(self, type_code):
        """获取题型中文名称"""
        type_names = {
            'single': '单选题',
            'multiple': '多选题',
            'judge': '判断题',
            'fill': '填空题',
            'essay': '简答题'
        }
        return type_names.get(type_code, type_code)
    
    def reset_ui(self):
        """重置UI状态"""
        self.test_button.config(state='normal')
        self.convert_button.config(state='normal')
        self.cancel_button.config(state='disabled')
        self.is_converting = False
    
    def open_output_directory(self):
        """打开输出目录"""
        if hasattr(self, 'output_file_path') and os.path.exists(self.output_file_path):
            # 获取文件所在目录
            output_dir = os.path.dirname(self.output_file_path)
            try:
                # Windows
                if sys.platform == 'win32':
                    os.startfile(output_dir)
                # macOS
                elif sys.platform == 'darwin':
                    os.system(f'open "{output_dir}"')
                # Linux
                else:
                    os.system(f'xdg-open "{output_dir}"')
            except Exception as e:
                messagebox.showerror("错误", f"无法打开目录: {str(e)}")
        else:
            messagebox.showwarning("警告", "输出文件不存在")
    
    def open_output_file(self):
        """打开输出文件（已弃用，保留用于兼容）"""
        self.open_output_directory()
    
    def convert_file(self, source_file, output_dir, file_format):
        """执行文件转换（在后台线程中运行）"""
        try:
            self.log_message(f"开始转换文件: {source_file}")
            self.log_message(f"文件格式: {file_format}")
            
            # 解析文件
            if file_format == 'excel':
                questions = self.parse_excel_file(source_file)
            elif file_format == 'word':
                questions = self.parse_word_file(source_file)
            else:
                raise ValueError(f"不支持的文件格式: {file_format}")
            
            if not questions:
                raise ValueError("未能从文件中提取到任何题目")
            
            self.log_message(f"成功解析 {len(questions)} 道题目")
            
            # 转换为JSON格式
            self.save_as_json(questions, source_file, output_dir)
            
        except InterruptedError as e:
            # 用户取消
            self.log_message(str(e))
            self.update_status("已取消")
            self.window.after(0, self.reset_ui)
            
        except FileNotFoundError as e:
            # 文件不存在
            error_msg = f"文件不存在: {str(e)}"
            self.log_message(f"错误: {error_msg}")
            self.update_status("转换失败")
            self.window.after(0, lambda: self.show_error_with_retry(
                "文件错误",
                error_msg,
                "请检查文件路径是否正确"
            ))
            
        except PermissionError as e:
            # 权限错误
            error_msg = f"权限错误: {str(e)}"
            self.log_message(f"错误: {error_msg}")
            self.update_status("转换失败")
            self.window.after(0, lambda: self.show_error_with_retry(
                "权限错误",
                error_msg,
                "请检查文件是否被其他程序占用，或者您是否有读取权限"
            ))
            
        except ValueError as e:
            # 数据验证错误
            error_msg = str(e)
            self.log_message(f"错误: {error_msg}")
            self.update_status("转换失败")
            self.window.after(0, lambda: self.show_error_with_retry(
                "数据错误",
                error_msg,
                "请检查文件格式是否正确，确保包含有效的题目数据"
            ))
            
        except Exception as e:
            # 其他错误
            error_msg = f"转换失败: {str(e)}"
            self.log_message(f"错误: {error_msg}")
            self.update_status("转换失败")
            
            # 获取详细的错误堆栈
            import traceback
            error_details = traceback.format_exc()
            self.log_message(f"详细错误信息:\n{error_details}")
            
            self.window.after(0, lambda: self.show_error_with_retry(
                "转换失败",
                error_msg,
                "请查看日志了解详细错误信息"
            ))
    
    def show_error_with_retry(self, title, message, suggestion):
        """显示错误消息并提供重试选项"""
        full_message = f"{message}\n\n{suggestion}\n\n是否重试？"
        
        if messagebox.askyesno(title, full_message):
            # 用户选择重试
            self.reset_ui()
            self.log_message("用户选择重试")
        else:
            # 用户选择不重试
            self.reset_ui()
    
    def parse_excel_file(self, file_path):
        """解析Excel文件"""
        self.log_message("正在解析Excel文件...")
        self.update_progress(10, "正在读取Excel文件...")
        
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        self.log_message(f"工作表: {sheet.title}, 行数: {sheet.max_row}")
        
        # 检测列结构
        columns = self.detect_excel_columns(sheet)
        if not columns:
            raise ValueError("无法识别Excel文件结构，请确保文件包含题目、选项、答案等列")
        
        self.log_message(f"检测到的列结构: {columns}")
        self.update_progress(20, "正在提取题目...")
        
        questions = []
        start_row = 2 if (columns['type'] >= 0 or columns['question'] >= 0) else 1
        total_rows = sheet.max_row - start_row + 1
        
        for row_idx in range(start_row, sheet.max_row + 1):
            if not self.is_converting:
                raise InterruptedError("用户取消了转换")
            
            row = [cell.value for cell in sheet[row_idx]]
            if not row or all(cell is None for cell in row):
                continue
            
            # 提取数据
            type_value = row[columns['type']] if columns['type'] >= 0 else ''
            question_value = row[columns['question']] if columns['question'] >= 0 else row[0]
            options_value = row[columns['options']] if columns['options'] >= 0 else (row[1] if len(row) > 1 else '')
            answer_value = row[columns['answer']] if columns['answer'] >= 0 else (row[2] if len(row) > 2 else '')
            explanation_value = row[columns['explanation']] if columns['explanation'] >= 0 else (row[3] if len(row) > 3 else '')
            
            if not question_value:
                continue
            
            # 标准化数据
            question_type = self.normalize_question_type(type_value)
            question = self.clean_text(question_value)
            options = self.parse_options(options_value, question_type)
            answer = self.normalize_answer(answer_value, question_type)
            explanation = self.clean_text(explanation_value) if explanation_value else ''
            
            # 验证题目
            validation_error = self.validate_question(question, question_type, options, answer)
            if validation_error:
                self.log_message(f"警告: 第 {row_idx} 行 - {validation_error}")
                continue
            
            questions.append({
                'content': question,
                'type': question_type,
                'options': options,
                'answer': answer,
                'explanation': explanation
            })
            
            # 更新进度
            progress = 20 + int((row_idx - start_row + 1) / total_rows * 60)
            self.update_progress(progress, f"已处理: {len(questions)}/{row_idx - start_row + 1} 题")
        
        self.update_progress(80, f"解析完成，共 {len(questions)} 道题目")
        return questions
    
    def compress_image(self, image_data, max_width=800, quality=85):
        """
        压缩图片以减小Base64编码后的大小
        
        参数:
            image_data: 原始图片二进制数据
            max_width: 最大宽度（像素），默认800
            quality: JPEG质量（1-100），默认85
        
        返回:
            压缩后的图片二进制数据
        """
        # 检查PIL是否可用
        if not PIL_AVAILABLE:
            self.log_message("警告: PIL库不可用，无法压缩图片，使用原始图片")
            return image_data
        
        try:
            from PIL import Image
            import io
            
            original_size = len(image_data)
            self.log_message(f"开始压缩图片，原始大小: {original_size/1024:.1f}KB")
            
            # 从二进制数据加载图片
            img = Image.open(io.BytesIO(image_data))
            self.log_message(f"图片信息: {img.width}x{img.height}, 模式: {img.mode}")
            
            # 转换RGBA为RGB（JPEG不支持透明度）
            if img.mode in ('RGBA', 'LA', 'P'):
                self.log_message(f"转换图片模式: {img.mode} → RGB")
                # 创建白色背景
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                self.log_message(f"转换图片模式: {img.mode} → RGB")
                img = img.convert('RGB')
            
            # 计算新尺寸（保持宽高比）
            if img.width > max_width:
                ratio = max_width / img.width
                new_height = int(img.height * ratio)
                old_size = f"{img.width}x{img.height}"
                img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
                self.log_message(f"调整图片尺寸: {old_size} → {img.width}x{img.height}")
            
            # 保存为JPEG格式（压缩）
            output = io.BytesIO()
            img.save(output, format='JPEG', quality=quality, optimize=True)
            compressed_data = output.getvalue()
            
            # 计算压缩比
            compressed_size = len(compressed_data)
            ratio = (1 - compressed_size / original_size) * 100
            self.log_message(f"✓ 压缩完成: {original_size/1024:.1f}KB → {compressed_size/1024:.1f}KB (减少 {ratio:.1f}%)")
            
            return compressed_data
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log_message(f"✗ 图片压缩失败，使用原始图片")
            self.log_message(f"错误详情: {str(e)}")
            self.log_message(f"堆栈信息:\n{error_detail}")
            return image_data
    
    def parse_word_file(self, file_path):
        """解析Word文件（支持图片提取和更好的题型识别）"""
        self.log_message("正在解析Word文件...")
        self.update_progress(10, "正在读取Word文件...")
        
        doc = Document(file_path)
        
        # 图片统计
        total_images = 0
        
        # 提取所有段落文本和图片
        lines = []
        for para in doc.paragraphs:
            text = para.text.strip()
            
            # 提取段落中的所有图片
            para_images = []
            for run in para.runs:
                # 检查run中是否有图片
                drawing_elements = run._element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}drawing')
                
                for drawing in drawing_elements:
                    # 查找图片的blip元素（包含图片引用）
                    blips = drawing.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
                    
                    for blip in blips:
                        # 获取图片的关系ID
                        embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                        
                        if embed_id:
                            try:
                                total_images += 1
                                self.log_message(f"\n{'='*60}")
                                self.log_message(f"处理第 {total_images} 张图片 (段落 {len(lines)+1})")
                                self.log_message(f"{'='*60}")
                                
                                # 通过关系ID获取图片数据
                                image_part = run.part.related_parts[embed_id]
                                image_data = image_part.blob
                                
                                # 压缩图片
                                compressed_data = self.compress_image(image_data)
                                
                                import base64
                                b64_image = base64.b64encode(compressed_data).decode('utf-8')
                                b64_size = len(b64_image)
                                self.log_message(f"Base64编码后大小: {b64_size/1024:.1f}KB")
                                
                                # 压缩后统一使用JPEG格式
                                img_tag = f"<img src='data:image/jpeg;base64,{b64_image}' />"
                                para_images.append(img_tag)
                                
                            except Exception as e:
                                self.log_message(f"警告: 提取图片失败 - {str(e)}")
                                import traceback
                                self.log_message(traceback.format_exc())
            
            # 组合文本和图片
            if text or para_images:
                line_content = text
                if para_images:
                    # 将图片添加到文本后面
                    line_content += ''.join(para_images)
                lines.append(line_content)
        
        self.log_message(f"\n{'='*60}")
        self.log_message(f"文档解析完成")
        self.log_message(f"{'='*60}")
        self.log_message(f"共 {len(lines)} 个段落")
        self.log_message(f"共提取 {total_images} 张图片")
        self.log_message(f"{'='*60}\n")
        self.update_progress(20, "正在提取题目...")
        
        questions = []
        current_question = None
        options_list = []
        
        for i, line in enumerate(lines):
            if not self.is_converting:
                raise InterruptedError("用户取消了转换")
            
            # 检测题目开始（支持多种格式）
            # 格式1: 1. 题目内容
            # 格式2: 1、题目内容
            # 格式3: 1．题目内容
            # 格式4: 1）题目内容
            # 格式5: 【单选题】题目内容
            # 格式6: 【复合题】题目内容
            # 格式7: 【问答题】题目内容
            # 格式8: (1)【问答题】答案内容（简答题答案格式）
            question_match = re.match(r'^(\d+)[.、．)\s]+(.+)', line)
            type_match = re.match(r'^【(单选题|多选题|判断题|填空题|简答题|问答题|复合题)】(.+)', line)
            answer_with_type_match = re.match(r'^\((\d+)\)【(问答题|简答题)】(.+)', line)
            
            # 检测简答题答案格式：(1)【问答题】答案内容
            if answer_with_type_match and current_question and current_question['type'] in ['essay', 'short_answer']:
                # 这是上一题的答案
                answer_text = answer_with_type_match.group(3).strip()
                current_question['answer'] = answer_text
                continue
            
            if question_match or type_match:
                # 保存上一题
                if current_question and current_question['content']:
                    if options_list:
                        current_question['options'] = options_list
                    
                    # 清理临时标记
                    if '_collecting_answer' in current_question:
                        del current_question['_collecting_answer']
                    
                    # 智能判断题型（如果没有明确的题型标记）
                    if current_question['type'] == 'single':  # 默认类型，需要智能判断
                        current_question['type'] = self.detect_question_type(
                            current_question['content'],
                            options_list,
                            current_question['answer']
                        )
                    
                    questions.append(current_question)
                
                # 开始新题
                if type_match:
                    # 从题型标记中提取题型
                    type_str = type_match.group(1)
                    content = type_match.group(2).strip()
                    q_type = self.normalize_question_type(type_str)
                else:
                    content = question_match.group(2).strip()
                    q_type = 'single'  # 默认单选题
                
                current_question = {
                    'content': content,
                    'type': q_type,
                    'options': [],
                    'answer': '',
                    'explanation': ''
                }
                options_list = []
                continue
            
            if not current_question:
                continue
            
            # 检测选项（支持多种格式）
            # 格式1: A. 选项内容
            # 格式2: A、选项内容
            # 格式3: A．选项内容
            # 格式4: A）选项内容
            option_match = re.match(r'^([A-Z])[.、．)\s]+(.+)', line)
            if option_match:
                options_list.append(option_match.group(2).strip())
                continue
            
            # 检测答案（支持多种格式）
            # 格式1: 答案：A（答案在同一行）
            # 格式2: 答案：（答案在后续行，直到遇到"解析："或下一题）
            # 支持的答案关键字：答案、正确答案、参考答案
            answer_match = re.match(r'^(?:答案|正确答案|参考答案)[:：]\s*(.*)$', line, re.IGNORECASE)
            if answer_match:
                answer_text = answer_match.group(1).strip()
                if answer_text:
                    # 答案在同一行
                    current_question['answer'] = answer_text
                else:
                    # 答案在下一行，标记进入答案收集模式
                    current_question['_collecting_answer'] = True
                    current_question['answer'] = ''  # 初始化为空字符串
                continue
            
            # 检测解析（解析标记会结束答案收集）
            # 支持的解析关键字：解析、答案解析、说明、提示
            explanation_match = re.match(r'^(?:解析|答案解析|说明|提示)[:：]\s*(.*)$', line, re.IGNORECASE)
            if explanation_match:
                # 结束答案收集
                if current_question and current_question.get('_collecting_answer'):
                    current_question['_collecting_answer'] = False
                
                explanation_text = explanation_match.group(1).strip()
                if explanation_text:
                    current_question['explanation'] = explanation_text
                continue
            
            # 如果正在收集答案（"答案："到"解析："之间的所有内容）
            if current_question and current_question.get('_collecting_answer'):
                # 所有非空行都是答案的一部分
                if line:
                    if current_question['answer']:
                        current_question['answer'] += '\n' + line
                    else:
                        current_question['answer'] = line
                continue
            
            # 题干延续（如果不是选项、答案、解析，就认为是题干的一部分）
            if current_question['content'] and not options_list and not current_question['answer']:
                # 不是以字母开头的行，认为是题干延续
                if not re.match(r'^[A-Z][.、．)]', line):
                    current_question['content'] += ' ' + line
            
            # 更新进度
            progress = 20 + int((i + 1) / len(lines) * 60)
            self.update_progress(progress, f"正在处理第 {i + 1}/{len(lines)} 段")
        
        # 保存最后一题
        if current_question and current_question['content']:
            if options_list:
                current_question['options'] = options_list
            
            # 清理临时标记
            if '_collecting_answer' in current_question:
                del current_question['_collecting_answer']
            
            # 智能判断题型（如果没有明确的题型标记）
            if current_question['type'] == 'single':  # 默认类型，需要智能判断
                current_question['type'] = self.detect_question_type(
                    current_question['content'],
                    options_list,
                    current_question['answer']
                )
            
            questions.append(current_question)
        
        self.update_progress(80, f"解析完成，共 {len(questions)} 道题目")
        return questions
    
    def detect_question_type(self, content, options, answer):
        """智能检测题型"""
        has_options = options and len(options) > 0
        
        # 1. 检测填空题：题目中包含下划线（但如果有选项，则不是填空题）
        if not has_options:
            # 检测连续的下划线（至少3个）
            if '___' in content or '____' in content or '______' in content:
                if self.debug_mode:
                    self.log_message("")
                    self.log_message("=" * 50)
                    self.log_message(f"[题型识别] 题目: {content[:80]}...")
                    self.log_message(f"[题型识别] ✓ 检测到填空标记（无选项）")
                    self.log_message(f"[题型识别] → 识别为【填空题】")
                return 'fill'
        
        # 2. 检测简答题/问答题：没有选项或选项为空，且答案较长，或包含特定关键词
        if not has_options:
            # 检查题目内容是否包含问答题特征
            keywords = ['阅读材料', '根据材料', '结合所学', '分析', '概括', '归纳', '论述', '简述', '说明', '回答问题', '完成任务']
            matched_keywords = [kw for kw in keywords if kw in content]
            
            if matched_keywords:
                if self.debug_mode:
                    self.log_message("")
                    self.log_message("=" * 50)
                    self.log_message(f"[题型识别] 题目: {content[:80]}...")
                    self.log_message(f"[题型识别] ✓ 匹配到简答题关键词: {matched_keywords}")
                    self.log_message(f"[题型识别] → 识别为【简答题】")
                return 'essay'
            
            # 答案较长也认为是简答题
            if answer and len(str(answer).strip()) > 10:
                if self.debug_mode:
                    self.log_message("")
                    self.log_message("=" * 50)
                    self.log_message(f"[题型识别] 题目: {content[:80]}...")
                    self.log_message(f"[题型识别] 答案: {str(answer)[:50]}...")
                    self.log_message(f"[题型识别] ✓ 答案长度 {len(str(answer).strip())} > 10")
                    self.log_message(f"[题型识别] → 识别为【简答题】")
                return 'essay'
        
        # 3. 检测判断题：只有2个选项，或答案是"正确/错误"
        if has_options and len(options) == 2:
            # 检查选项内容是否是判断题特征
            opt_text = ''.join(options).lower()
            if any(x in opt_text for x in ['正确', '错误', '对', '错', 'true', 'false']):
                if self.debug_mode:
                    self.log_message("")
                    self.log_message("=" * 50)
                    self.log_message(f"[题型识别] 题目: {content[:80]}...")
                    self.log_message(f"[题型识别] 选项: {options}")
                    self.log_message(f"[题型识别] ✓ 2个选项且包含判断题特征")
                    self.log_message(f"[题型识别] → 识别为【判断题】")
                return 'judge'
        
        if answer:
            ans_lower = str(answer).lower().strip()
            if ans_lower in ['正确', '错误', '对', '错', 'true', 'false', 't', 'f', 'a', 'b']:
                # 如果没有选项但答案是判断题答案，也认为是判断题
                if not has_options:
                    if self.debug_mode:
                        self.log_message("")
                        self.log_message("=" * 50)
                        self.log_message(f"[题型识别] 题目: {content[:80]}...")
                        self.log_message(f"[题型识别] 答案: {answer}")
                        self.log_message(f"[题型识别] ✓ 无选项但答案是判断题答案")
                        self.log_message(f"[题型识别] → 识别为【判断题】")
                    return 'judge'
        
        # 4. 检测多选题：答案包含多个字母
        if answer and len(re.findall(r'[A-Z]', str(answer).upper())) > 1:
            if self.debug_mode:
                self.log_message("")
                self.log_message("=" * 50)
                self.log_message(f"[题型识别] 题目: {content[:80]}...")
                self.log_message(f"[题型识别] 答案: {answer}")
                self.log_message(f"[题型识别] ✓ 答案包含多个字母")
                self.log_message(f"[题型识别] → 识别为【多选题】")
            return 'multiple'
        
        # 5. 默认单选题（不输出日志，减少干扰）
        return 'single'
    
    def detect_excel_columns(self, sheet):
        """智能检测Excel列结构"""
        if sheet.max_row == 0:
            return None
        
        # 读取第一行作为标题
        header_row = [cell.value for cell in sheet[1]]
        
        columns = {
            'type': -1,
            'question': -1,
            'options': -1,
            'answer': -1,
            'explanation': -1
        }
        
        # 检测列
        for idx, header in enumerate(header_row):
            if not header:
                continue
            h = str(header).strip().lower()
            
            if '题型' in h or 'type' in h or h == '类型':
                columns['type'] = idx
            elif '题干' in h or '题目' in h or 'question' in h or h == '内容':
                columns['question'] = idx
            elif '选项' in h or 'option' in h:
                columns['options'] = idx
            elif '答案' in h or 'answer' in h or h == '正确答案':
                columns['answer'] = idx
            elif '解析' in h or 'explanation' in h or '说明' in h:
                columns['explanation'] = idx
        
        # 如果没有检测到标题，按位置推断
        if columns['question'] == -1:
            if len(header_row) >= 5:
                columns = {'type': 0, 'question': 1, 'options': 2, 'answer': 3, 'explanation': 4}
            elif len(header_row) >= 4:
                columns = {'type': -1, 'question': 0, 'options': 1, 'answer': 2, 'explanation': 3}
        
        return columns
    
    def clean_text(self, text):
        """清理文本中的HTML标签和多余空白"""
        if not text:
            return ''
        
        text = str(text)
        # 移除HTML标签
        text = re.sub(r'<[^>]+>', '', text)
        # 清理多余的空白字符
        text = re.sub(r'\s+', ' ', text)
        return text.strip()
    
    def normalize_question_type(self, type_str):
        """标准化题型"""
        if not type_str:
            return 'single'
        
        type_str = str(type_str).strip().lower()
        
        if '单选' in type_str or type_str in ['single', 'a', '1']:
            return 'single'
        if '多选' in type_str or type_str in ['multiple', 'b', '2']:
            return 'multiple'
        if '判断' in type_str or type_str in ['judge', 'c', '3']:
            return 'judge'
        if '填空' in type_str or type_str in ['fill', 'fill_in_blank', 'd', '4']:
            return 'fill'
        if '简答' in type_str or '问答' in type_str or '复合' in type_str or type_str in ['essay', 'short_answer', 'e', '5']:
            return 'essay'
        
        return 'single'
    
    def normalize_answer(self, answer, question_type):
        """标准化答案格式"""
        if not answer:
            return ''
        
        answer_str = str(answer).strip().upper()
        
        if question_type == 'judge':
            # 判断题答案转换
            if any(x in answer_str for x in ['正确', '对', '√', 'T', 'TRUE']) or answer_str == 'A' or answer_str == '1':
                return 'A'
            if any(x in answer_str for x in ['错误', '错', '×', 'F', 'FALSE']) or answer_str == 'B' or answer_str == '0':
                return 'B'
        
        # 移除所有非字母字符
        return re.sub(r'[^A-Z]', '', answer_str)
    
    def parse_options(self, options_value, question_type):
        """解析选项"""
        if question_type == 'judge':
            return ['正确', '错误']
        
        if not options_value:
            return []
        
        options_str = str(options_value).strip()
        options_list = []
        
        # 尝试不同的分隔符
        if '|' in options_str:
            options_list = options_str.split('|')
        elif '\n' in options_str or '</p>' in options_str:
            if '</p>' in options_str:
                parts = re.split(r'</p>\s*<p>|</p>|<p>', options_str)
            else:
                parts = options_str.split('\n')
            
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                part = re.sub(r'<[^>]+>', '', part)
                part = part.strip()
                if part:
                    options_list.append(part)
        elif '；' in options_str:
            options_list = options_str.split('；')
        elif ';' in options_str:
            options_list = options_str.split(';')
        else:
            clean_str = re.sub(r'<[^>]+>', '', options_str)
            matches = re.findall(r'[A-Z][.、．]\s*[^A-Z.、．]+', clean_str)
            if matches:
                options_list = matches
            else:
                options_list = [options_str]
        
        # 清理选项
        cleaned_options = []
        for opt in options_list:
            opt = opt.strip()
            if not opt:
                continue
            opt = re.sub(r'<[^>]+>', '', opt)
            opt = re.sub(r'^[A-Z][.、．]\s*', '', opt)
            opt = re.sub(r'^[①②③④⑤⑥⑦⑧]\s*', '', opt)
            opt = re.sub(r'^\d+[.、．]\s*', '', opt)
            opt = re.sub(r'^[（(]\s*[A-Z]\s*[)）]\s*', '', opt)
            opt = re.sub(r'\s+', ' ', opt).strip()
            if opt:
                cleaned_options.append(opt)
        
        return cleaned_options
    
    def validate_question(self, question, question_type, options, answer):
        """验证题目数据"""
        if not question:
            return "题目内容不能为空"
        
        if not answer:
            return "答案不能为空"
        
        if question_type in ['single', 'multiple']:
            if not options or len(options) < 2:
                return f"选择题至少需要2个选项（当前: {len(options) if options else 0}个）"
        
        return None
    
    def save_as_json(self, questions, source_file, output_dir):
        """将题目保存为JSON格式"""
        self.log_message("正在生成JSON文件...")
        self.update_progress(85, "正在生成JSON...")
        
        # 生成输出文件名
        source_name = Path(source_file).stem
        output_filename = f"{source_name}_converted.json"
        output_path = os.path.join(output_dir, output_filename)
        
        # 构建JSON数据结构
        json_data = {
            "metadata": {
                "version": "2.0",
                "createdAt": datetime.now().isoformat(),
                "totalQuestions": len(questions),
                "source": "python-gui-converter",
                "sourceFile": os.path.basename(source_file)
            },
            "questions": []
        }
        
        # 转换题目格式
        for q in questions:
            question_data = {
                "content": q['content'],
                "type": q['type'],
                "answer": q['answer'],
                "explanation": q.get('explanation', ''),
                "difficulty": 1,
                "tags": []
            }
            
            # 添加选项（如果有）
            if q.get('options'):
                question_data['options'] = q['options']
            
            # 简答题需要额外添加 referenceAnswer 字段（用于导入到数据库）
            if q['type'] == 'essay':
                question_data['referenceAnswer'] = q['answer']
            
            json_data['questions'].append(question_data)
        
        # 保存JSON文件
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            self.output_file_path = output_path
            self.log_message(f"成功保存到: {output_path}")
            self.update_progress(100, f"转换完成！共 {len(questions)} 道题目")
            self.update_status("转换成功")
            
            # 显示成功消息
            self.window.after(0, lambda: self.show_success_message(len(questions), output_path))
            
            # 启用打开文件按钮
            self.window.after(0, lambda: self.open_button.config(state='normal'))
            
        except Exception as e:
            raise Exception(f"保存JSON文件失败: {str(e)}")
        finally:
            self.window.after(0, self.reset_ui)
    
    def show_success_message(self, question_count, output_path):
        """显示转换成功消息"""
        message = f"转换成功！\n\n"
        message += f"题目数量: {question_count}\n"
        message += f"输出文件: {output_path}\n\n"
        message += "是否立即打开输出目录？"
        
        if messagebox.askyesno("转换成功", message):
            self.open_output_directory()
    
    def run(self):
        """运行GUI应用"""
        self.window.mainloop()


def main():
    """主函数"""
    app = QuestionBankConverterGUI()
    app.run()


if __name__ == '__main__':
    main()
