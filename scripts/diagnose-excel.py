#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 文件诊断脚本
检查原始题库文件中的问题
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误: 未安装 openpyxl")
    print("安装命令: pip install openpyxl")
    sys.exit(1)


def diagnose_excel(file_path):
    """诊断 Excel 文件"""
    print(f"\n正在诊断: {file_path}")
    print("=" * 60)
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # 读取标题行
        header_row = [cell.value for cell in sheet[1]]
        print(f"\n标题行: {header_row}")
        
        # 检测列结构
        columns = {}
        for idx, header in enumerate(header_row):
            if not header:
                continue
            h = str(header).strip().lower()
            
            if '题型' in h or 'type' in h:
                columns['type'] = idx
            elif '题干' in h or '题目' in h or 'question' in h:
                columns['question'] = idx
            elif '选项' in h or 'option' in h:
                columns['options'] = idx
            elif '答案' in h or 'answer' in h:
                columns['answer'] = idx
            elif '解析' in h or 'explanation' in h:
                columns['explanation'] = idx
        
        print(f"\n检测到的列结构:")
        for key, idx in columns.items():
            print(f"  {key}: 第 {idx + 1} 列 ({header_row[idx]})")
        
        # 检查数据
        print(f"\n数据检查:")
        print(f"  总行数: {sheet.max_row}")
        
        empty_options = []
        empty_questions = []
        
        for row_idx in range(2, min(sheet.max_row + 1, 50)):  # 只检查前 50 行
            row = [cell.value for cell in sheet[row_idx]]
            
            question_value = row[columns['question']] if 'question' in columns else row[0]
            options_value = row[columns['options']] if 'options' in columns else (row[1] if len(row) > 1 else '')
            
            if not question_value or str(question_value).strip() == '':
                empty_questions.append(row_idx)
            
            if not options_value or str(options_value).strip() == '':
                empty_options.append(row_idx)
        
        if empty_questions:
            print(f"\n⚠️  发现 {len(empty_questions)} 行题干为空:")
            print(f"  行号: {empty_options[:10]}")
            if len(empty_questions) > 10:
                print(f"  ...(还有 {len(empty_questions) - 10} 行)")
        
        if empty_options:
            print(f"\n⚠️  发现 {len(empty_options)} 行选项为空:")
            print(f"  行号: {empty_options[:10]}")
            if len(empty_options) > 10:
                print(f"  ...(还有 {len(empty_options) - 10} 行)")
            
            # 显示前几个空选项的题目
            print(f"\n  示例（前 5 个）:")
            for row_idx in empty_options[:5]:
                row = [cell.value for cell in sheet[row_idx]]
                question = row[columns['question']] if 'question' in columns else row[0]
                print(f"    第 {row_idx} 行: {str(question)[:60]}...")
        
        # 检查选项格式
        print(f"\n选项格式检查（前 10 行）:")
        for row_idx in range(2, min(12, sheet.max_row + 1)):
            row = [cell.value for cell in sheet[row_idx]]
            options_value = row[columns['options']] if 'options' in columns else (row[1] if len(row) > 1 else '')
            
            if options_value:
                options_str = str(options_value)
                has_pipe = '|' in options_str
                has_newline = '\n' in options_str
                has_semicolon = ';' in options_str or '；' in options_str
                
                print(f"  第 {row_idx} 行:")
                print(f"    内容: {options_str[:80]}...")
                print(f"    分隔符: ", end='')
                if has_pipe:
                    print("✓ 竖线(|)", end=' ')
                if has_newline:
                    print("✓ 换行", end=' ')
                if has_semicolon:
                    print("✓ 分号", end=' ')
                if not (has_pipe or has_newline or has_semicolon):
                    print("❌ 未检测到分隔符")
                else:
                    print()
        
        print("\n" + "=" * 60)
        
    except Exception as e:
        print(f"❌ 诊断失败: {str(e)}")
        import traceback
        traceback.print_exc()


def main():
    """主函数"""
    print("=" * 60)
    print("Excel 文件诊断工具")
    print("=" * 60)
    
    # 扫描文件
    current_dir = Path('.')
    files = list(current_dir.glob('原始题库*.xlsx'))
    
    if not files:
        print("\n❌ 未找到原始题库文件")
        print("请将文件命名为 '原始题库*.xlsx'")
        return
    
    print(f"\n找到 {len(files)} 个文件:")
    for f in files:
        print(f"  - {f.name}")
    
    # 诊断每个文件
    for file_path in files:
        diagnose_excel(str(file_path))
    
    print("\n✅ 诊断完成")
    print("\n建议:")
    print("  1. 检查原始 Excel 文件中选项列是否有数据")
    print("  2. 确保选项使用正确的分隔符（推荐使用 | ）")
    print("  3. 对于选项为空的题目，请在 Excel 中补充完整")


if __name__ == '__main__':
    main()
