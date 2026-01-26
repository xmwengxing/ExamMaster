#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
题目格式转换脚本 - Python版本
将原始题库文件（Excel/Word）转换为系统要求的CSV格式

使用方法：
1. 安装依赖：pip install openpyxl python-docx
2. 运行脚本：python scripts/convert-questions.py
3. 指定章节：python scripts/convert-questions.py --chapter "九上第一单元"
"""

import os
import re
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: 未安装 openpyxl，无法处理Excel文件")
    print("安装命令: pip install openpyxl")

try:
    from docx import Document
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False
    print("警告: 未安装 python-docx，无法处理Word文件")
    print("安装命令: pip install python-docx")

# 全局变量：用户指定的章节名称
CUSTOM_CHAPTER = None


def clean_html_tags(text):
    """清理 HTML 标签和多余空白"""
    if not text:
        return ''
    
    text = str(text)
    # 移除 HTML 标签
    text = re.sub(r'<[^>]+>', '', text)
    # 清理多余的空白字符（包括换行符）
    text = re.sub(r'\s+', ' ', text)
    # 清理首尾空白
    return text.strip()


def normalize_question_type(type_str):
    """标准化题型"""
    if not type_str:
        return 'SINGLE'
    
    type_str = str(type_str).strip().upper()
    
    if '单选' in type_str or type_str in ['SINGLE', 'A', '1']:
        return 'SINGLE'
    if '多选' in type_str or type_str in ['MULTIPLE', 'B', '2']:
        return 'MULTIPLE'
    if '判断' in type_str or type_str in ['JUDGE', 'C', '3']:
        return 'JUDGE'
    
    return 'SINGLE'


def normalize_answer(answer, question_type):
    """标准化答案格式"""
    if not answer:
        return ''
    
    answer_str = str(answer).strip().upper()
    
    if question_type == 'JUDGE':
        # 判断题答案转换
        if any(x in answer_str for x in ['正确', '对', '√', 'T', 'TRUE']) or answer_str == 'A' or answer_str == '1':
            return 'A'
        if any(x in answer_str for x in ['错误', '错', '×', 'F', 'FALSE']) or answer_str == 'B' or answer_str == '0':
            return 'B'
    
    # 移除所有非字母字符
    return re.sub(r'[^A-Z]', '', answer_str)


def normalize_options(options, question_type):
    """处理选项：统一为 A|B|C|D 格式"""
    if question_type == 'JUDGE':
        return ''
    
    if not options:
        return ''
    
    options_str = str(options).strip()
    
    # 先尝试按分隔符分割（在清理 HTML 之前）
    options_list = []
    
    if '|' in options_str:
        # 已经是标准格式
        options_list = options_str.split('|')
    elif '\n' in options_str or '</p>' in options_str:
        # 包含换行符或 HTML 标签的格式
        # 先按 </p> 或换行符分割
        if '</p>' in options_str:
            # HTML 格式：<p>A. 选项1</p><p>B. 选项2</p>
            parts = re.split(r'</p>\s*<p>|</p>|<p>', options_str)
        else:
            # 纯换行符格式
            parts = options_str.split('\n')
        
        # 清理每个部分
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # 移除 HTML 标签
            part = re.sub(r'<[^>]+>', '', part)
            part = part.strip()
            if part:
                options_list.append(part)
    elif '；' in options_str:
        options_list = options_str.split('；')
    elif ';' in options_str:
        options_list = options_str.split(';')
    else:
        # 尝试匹配 A. B. C. D. 格式（先清理 HTML）
        clean_str = re.sub(r'<[^>]+>', '', options_str)
        matches = re.findall(r'[A-Z][.、．]\s*[^A-Z.、．]+', clean_str)
        if matches:
            options_list = matches
        else:
            options_list = [options_str]
    
    # 清理选项：移除序号和多余空白
    cleaned_options = []
    for opt in options_list:
        opt = opt.strip()
        if not opt:
            continue
        # 移除 HTML 标签（再次清理，确保干净）
        opt = re.sub(r'<[^>]+>', '', opt)
        # 移除各种序号格式
        opt = re.sub(r'^[A-Z][.、．]\s*', '', opt)
        opt = re.sub(r'^[①②③④⑤⑥⑦⑧]\s*', '', opt)
        opt = re.sub(r'^\d+[.、．]\s*', '', opt)
        opt = re.sub(r'^[（(]\s*[A-Z]\s*[)）]\s*', '', opt)
        # 清理多余空白
        opt = re.sub(r'\s+', ' ', opt).strip()
        if opt:
            cleaned_options.append(opt)
    
    return '|'.join(cleaned_options)


def detect_excel_columns(sheet):
    """智能检测Excel列结构"""
    if sheet.max_row == 0:
        return None
    
    # 读取第一行作为标题
    header_row = [cell.value for cell in sheet[1]]
    
    columns = {
        'type': -1,
        'question': -1,
        'options': -1,
        'answer': -1,
        'explanation': -1
    }
    
    # 检测列
    for idx, header in enumerate(header_row):
        if not header:
            continue
        h = str(header).strip().lower()
        
        if '题型' in h or 'type' in h or h == '类型':
            columns['type'] = idx
        elif '题干' in h or '题目' in h or 'question' in h or h == '内容':
            columns['question'] = idx
        elif '选项' in h or 'option' in h:
            columns['options'] = idx
        elif '答案' in h or 'answer' in h or h == '正确答案':
            columns['answer'] = idx
        elif '解析' in h or 'explanation' in h or '说明' in h:
            columns['explanation'] = idx
    
    # 如果没有检测到标题，按位置推断
    if columns['question'] == -1:
        if len(header_row) >= 5:
            columns = {'type': 0, 'question': 1, 'options': 2, 'answer': 3, 'explanation': 4}
        elif len(header_row) >= 4:
            columns = {'type': -1, 'question': 0, 'options': 1, 'answer': 2, 'explanation': 3}
    
    print(f"  检测到的列结构: {columns}")
    return columns


def convert_excel_file(file_path):
    """转换Excel文件"""
    print(f"正在处理: {file_path}")
    
    if not EXCEL_AVAILABLE:
        print("  ✗ 跳过: 未安装 openpyxl")
        return []
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # 检测列结构
        columns = detect_excel_columns(sheet)
        if not columns:
            print("  ⚠ 无法识别文件结构")
            return []
        
        questions = []
        start_row = 2 if (columns['type'] >= 0 or columns['question'] >= 0) else 1
        
        for row_idx in range(start_row, sheet.max_row + 1):
            row = [cell.value for cell in sheet[row_idx]]
            if not row or all(cell is None for cell in row):
                continue
            
            # 提取数据
            type_value = row[columns['type']] if columns['type'] >= 0 else ''
            question_value = row[columns['question']] if columns['question'] >= 0 else row[0]
            options_value = row[columns['options']] if columns['options'] >= 0 else (row[1] if len(row) > 1 else '')
            answer_value = row[columns['answer']] if columns['answer'] >= 0 else (row[2] if len(row) > 2 else '')
            explanation_value = row[columns['explanation']] if columns['explanation'] >= 0 else (row[3] if len(row) > 3 else '')
            
            if not question_value:
                continue
            
            question_type = normalize_question_type(type_value)
            question = clean_html_tags(question_value)
            options = normalize_options(options_value, question_type)
            answer = normalize_answer(answer_value, question_type)
            explanation = clean_html_tags(explanation_value) if explanation_value else ''
            
            # 验证：选择题必须有选项
            if question_type in ['SINGLE', 'MULTIPLE'] and not options:
                print(f"  ⚠ 跳过第 {row_idx} 行：选择题缺少选项")
                continue
            
            # 验证：选择题至少需要 2 个选项
            if question_type in ['SINGLE', 'MULTIPLE']:
                option_list = options.split('|') if options else []
                if len(option_list) < 2:
                    print(f"  ⚠ 跳过第 {row_idx} 行：选项不足（需要至少2个，实际{len(option_list)}个）")
                    continue
            
            if question:
                questions.append({
                    'type': question_type,
                    'question': question,
                    'options': options,
                    'answer': answer,
                    'explanation': explanation
                })
        
        print(f"  ✓ 成功转换 {len(questions)} 道题目")
        return questions
        
    except Exception as e:
        print(f"  ✗ 处理失败: {str(e)}")
        return []


def convert_word_file(file_path):
    """转换Word文档"""
    print(f"正在处理: {file_path}")
    
    if not WORD_AVAILABLE:
        print("  ✗ 跳过: 未安装 python-docx")
        return []
    
    try:
        doc = Document(file_path)
        
        # 提取所有段落文本
        lines = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                lines.append(text)
        
        questions = []
        current_question = None
        options_list = []
        
        for line in lines:
            # 检测题目开始
            question_match = re.match(r'^(\d+)[.、．)\s]+(.+)', line)
            if question_match:
                # 保存上一题
                if current_question and current_question['question']:
                    if options_list:
                        current_question['options'] = '|'.join(options_list)
                    questions.append(current_question)
                
                # 开始新题
                current_question = {
                    'type': 'SINGLE',
                    'question': question_match.group(2).strip(),
                    'options': '',
                    'answer': '',
                    'explanation': ''
                }
                options_list = []
                continue
            
            if not current_question:
                continue
            
            # 检测选项
            option_match = re.match(r'^([A-Z])[.、．)\s]+(.+)', line)
            if option_match:
                options_list.append(option_match.group(2).strip())
                continue
            
            # 检测答案
            answer_match = re.match(r'^(?:答案|正确答案|参考答案)[:：\s]+([A-Z]+|正确|错误|对|错)', line, re.IGNORECASE)
            if answer_match:
                current_question['answer'] = normalize_answer(answer_match.group(1), current_question['type'])
                
                # 根据答案判断题型
                if len(current_question['answer']) > 1:
                    current_question['type'] = 'MULTIPLE'
                elif not options_list or len(options_list) <= 2:
                    ans = answer_match.group(1).strip()
                    if ans in ['正确', '错误', '对', '错']:
                        current_question['type'] = 'JUDGE'
                        options_list = []
                continue
            
            # 检测解析
            explanation_match = re.match(r'^(?:解析|答案解析|说明)[:：\s]+(.+)', line, re.IGNORECASE)
            if explanation_match:
                current_question['explanation'] = explanation_match.group(1).strip()
                continue
            
            # 题干延续
            if current_question['question'] and not options_list and not current_question['answer'] and not re.match(r'^[A-Z][.、．)]', line):
                current_question['question'] += ' ' + line
        
        # 保存最后一题
        if current_question and current_question['question']:
            if options_list:
                current_question['options'] = '|'.join(options_list)
            questions.append(current_question)
        
        print(f"  ✓ 成功转换 {len(questions)} 道题目")
        return questions
        
    except Exception as e:
        print(f"  ✗ 处理失败: {str(e)}")
        return []


def generate_csv(questions, output_file):
    """生成CSV文件"""
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        
        # 写入标题（新格式：8个字段）
        writer.writerow([
            '题型(SINGLE/MULTIPLE/JUDGE/FILL_IN_BLANK/SHORT_ANSWER)', 
            '题干', 
            '选项(用|分隔)', 
            '答案', 
            '解析',
            '单元/章节',
            '填空配置(格式:blank1:答案1|答案2;blank2:答案3)',
            '简答参考答案'
        ])
        
        # 写入数据
        for q in questions:
            writer.writerow([
                q['type'],
                q['question'],
                q['options'],
                q['answer'],
                q['explanation'],
                q.get('chapter', ''),  # 单元/章节
                q.get('fillBlanks', ''),  # 填空配置
                q.get('shortAnswer', '')  # 简答参考答案
            ])
    
    print(f"\n✓ 已生成: {output_file}")


def generate_report(questions):
    """生成统计报告"""
    total = len(questions)
    single = sum(1 for q in questions if q['type'] == 'SINGLE')
    multiple = sum(1 for q in questions if q['type'] == 'MULTIPLE')
    judge = sum(1 for q in questions if q['type'] == 'JUDGE')
    with_explanation = sum(1 for q in questions if q['explanation'])
    
    print('\n' + '=' * 60)
    print('转换统计')
    print('=' * 60)
    print(f"总题目数: {total}")
    print(f"  - 单选题: {single} ({single/total*100:.1f}%)")
    print(f"  - 多选题: {multiple} ({multiple/total*100:.1f}%)")
    print(f"  - 判断题: {judge} ({judge/total*100:.1f}%)")
    print(f"包含解析: {with_explanation} ({with_explanation/total*100:.1f}%)")
    print('=' * 60)


def main():
    """主函数"""
    global CUSTOM_CHAPTER
    
    # 解析命令行参数
    if '--chapter' in sys.argv:
        try:
            chapter_index = sys.argv.index('--chapter')
            if chapter_index + 1 < len(sys.argv):
                CUSTOM_CHAPTER = sys.argv[chapter_index + 1]
        except (ValueError, IndexError):
            pass
    
    print('=' * 60)
    print('题目格式转换工具 - Python版本')
    print('=' * 60)
    print()
    
    if CUSTOM_CHAPTER:
        print(f'📌 已设置章节信息: {CUSTOM_CHAPTER}')
        print('   所有题目将统一使用此章节信息')
        print()
    
    # 检查依赖
    if not EXCEL_AVAILABLE and not WORD_AVAILABLE:
        print("错误: 未安装必要的依赖库")
        print("请运行: pip install openpyxl python-docx")
        return
    
    # 扫描文件
    current_dir = Path('.')
    files = []
    
    for pattern in ['原始题库*.xlsx', '原始题库*.xls', '原始题库*.docx']:
        files.extend(current_dir.glob(pattern))
    
    if not files:
        print('❌ 未找到原始题库文件')
        print('\n请将文件命名为"原始题库*.xlsx"或"原始题库*.docx"')
        print('例如: 原始题库1.xlsx、原始题库2.xlsx、原始题库3.docx')
        return
    
    print(f"找到 {len(files)} 个文件:")
    for f in files:
        print(f"  - {f.name}")
    print()
    
    # 处理每个文件
    all_questions = []
    
    for file_path in files:
        ext = file_path.suffix.lower()
        
        if ext in ['.xlsx', '.xls']:
            questions = convert_excel_file(str(file_path))
        elif ext == '.docx':
            questions = convert_word_file(str(file_path))
        else:
            continue
        
        # 如果用户指定了章节，覆盖所有题目的章节信息
        if CUSTOM_CHAPTER:
            for q in questions:
                q['chapter'] = CUSTOM_CHAPTER
        
        all_questions.extend(questions)
        print()
    
    if not all_questions:
        print('❌ 没有成功转换任何题目')
        return
    
    # 生成统计报告
    generate_report(all_questions)
    
    # 生成CSV文件
    print('\n正在生成CSV文件...\n')
    generate_csv(all_questions, '转换后的题目-合并.csv')
    
    # 按题型分类
    single_questions = [q for q in all_questions if q['type'] == 'SINGLE']
    multiple_questions = [q for q in all_questions if q['type'] == 'MULTIPLE']
    judge_questions = [q for q in all_questions if q['type'] == 'JUDGE']
    
    if single_questions:
        generate_csv(single_questions, '转换后的题目-单选题.csv')
    if multiple_questions:
        generate_csv(multiple_questions, '转换后的题目-多选题.csv')
    if judge_questions:
        generate_csv(judge_questions, '转换后的题目-判断题.csv')
    
    print('\n✅ 转换完成！')
    if CUSTOM_CHAPTER:
        print(f'\n📌 所有题目已设置章节: {CUSTOM_CHAPTER}')
    print('\n📝 提示: 生成的CSV文件已更新为新格式（8个字段）')
    print('        包含：题型、题干、选项、答案、解析、单元/章节、填空配置、简答参考答案')
    print('        请检查生成的CSV文件，确认格式正确后再导入系统。')


if __name__ == '__main__':
    main()
